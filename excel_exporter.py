"""
Excel export functionality for contest data
"""
import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Handles Excel file generation for contest data"""
    
    def __init__(self, database):
        self.database = database
        
    def _add_title_to_sheet(self, ws, title_text, start_row=1, max_col=8):
        """Add a formatted title to the worksheet"""
        # Merge cells for title
        end_col = get_column_letter(max_col)
        ws.merge_cells(f'A{start_row}:{end_col}{start_row}')
        
        # Create title cell
        title_cell = ws.cell(row=start_row, column=1, value=title_text)
        title_cell.font = Font(bold=True, size=18, name="Calibri", color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set title row height
        ws.row_dimensions[start_row].height = 30
        
        return start_row + 1
        
    def export_contest_participants(self, contest_id: int) -> str:
        """Export contest participants to Excel file"""
        try:
            # Get contest info
            contest = self.database.get_contest_by_id(contest_id)
            if not contest:
                raise ValueError("Contest not found")
                
            # Get participants with submissions
            participants = self.database.get_contest_participants_detailed(contest_id)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Ishtirokchilar"
            
            # Set headers
            headers = [
                "№", "Ism", "Username", "Telefon raqami", "Telegram ID", 
                "Qo'shilgan vaqti", "Yuborgan rasmlar", "Caption"
            ]
            
            # Enhanced styling for headers
            header_font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Create border style
            thick_border = Border(
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000')
            )
            
            # Set row height for header
            ws.row_dimensions[1].height = 25
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thick_border
                
            # Enhanced styling for data rows
            data_font = Font(size=12, name="Calibri")
            data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Create thin border for data cells
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Alternating row colors
            light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            # Add participant data with enhanced formatting
            for row, participant in enumerate(participants, 2):
                # Set row height
                ws.row_dimensions[row].height = 20
                
                # Choose fill color (alternating rows)
                fill_color = light_fill if row % 2 == 0 else white_fill
                
                # Add data with styling
                cells = [
                    (row, 1, row-1),  # №
                    (row, 2, participant['full_name']),  # Ism
                    (row, 3, f"@{participant['username']}" if participant['username'] else "N/A"),  # Username
                    (row, 4, participant.get('phone_number', 'Yuq')),  # Telefon raqami
                    (row, 5, participant['user_id']),  # Telegram ID
                    (row, 6, participant['joined_at']),  # Qo'shilgan vaqti
                    (row, 7, participant['submissions_count']),  # Yuborgan rasmlar
                    (row, 8, participant['latest_caption'] or "")  # Caption
                ]
                
                for r, c, value in cells:
                    cell = ws.cell(row=r, column=c, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    cell.fill = fill_color
                
            # Set optimal column widths
            column_widths = {
                'A': 8,   # №
                'B': 20,  # Ism
                'C': 18,  # Username
                'D': 18,  # Telefon raqami
                'E': 15,  # Telegram ID
                'F': 20,  # Qo'shilgan vaqti
                'G': 15,  # Yuborgan rasmlar
                'H': 30   # Caption
            }
            
            for column_letter, width in column_widths.items():
                ws.column_dimensions[column_letter].width = width
                
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"konkurs_{contest['title'].replace(' ', '_')}_{timestamp}.xlsx"
            
            # Ensure exports directory exists
            os.makedirs("exports", exist_ok=True)
            filepath = os.path.join("exports", filename)
            
            # Save file
            wb.save(filepath)
            
            logger.info(f"Excel file exported: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting Excel file: {e}")
            raise
    
    def export_users_with_referrals(self, bot_id: int) -> str:
        """Export users with their phone numbers and referral counts"""
        try:
            # Get bot info
            bot_info = self.database.get_bot_by_id(bot_id)
            if not bot_info:
                raise ValueError("Bot not found")
                
            # Get users data
            users_data = self.database.get_all_users_with_referrals_for_export(bot_id)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Foydalanuvchilar"
            
            # Set headers
            headers = [
                "№", "Ism", "Familiya", "Telefon raqami", "Referral orqali qo'shgan odamlar soni"
            ]
            
            # Enhanced styling for headers
            header_font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Create border style
            thick_border = Border(
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000')
            )
            
            # Set row height for header
            ws.row_dimensions[1].height = 25
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thick_border
                
            # Enhanced styling for data rows
            data_font = Font(size=12, name="Calibri")
            data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Create thin border for data cells
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Alternating row colors
            light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            # Add user data with enhanced formatting
            for row, user in enumerate(users_data, 2):
                # Set row height
                ws.row_dimensions[row].height = 20
                
                # Choose fill color (alternating rows)
                fill_color = light_fill if row % 2 == 0 else white_fill
                
                # Add data with styling
                cells = [
                    (row, 1, row-1),  # №
                    (row, 2, user.get('first_name', 'N/A')),  # Ism
                    (row, 3, user.get('last_name', 'N/A')),  # Familiya
                    (row, 4, user.get('phone_number', 'Berilmagan')),  # Telefon raqami
                    (row, 5, user.get('referral_count', 0))  # Referral soni
                ]
                
                for r, c, value in cells:
                    cell = ws.cell(row=r, column=c, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    cell.fill = fill_color
                
            # Set optimal column widths
            column_widths = {
                'A': 8,   # №
                'B': 20,  # Ism
                'C': 20,  # Familiya
                'D': 18,  # Telefon raqami
                'E': 25   # Referral orqali qo'shgan odamlar soni
            }
            
            for column_letter, width in column_widths.items():
                ws.column_dimensions[column_letter].width = width
                
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            bot_name = bot_info.get('name', 'bot').replace(' ', '_')
            filename = f"foydalanuvchilar_{bot_name}_{timestamp}.xlsx"
            
            # Ensure exports directory exists
            os.makedirs("exports", exist_ok=True)
            filepath = os.path.join("exports", filename)
            
            # Save file
            wb.save(filepath)
            
            logger.info(f"Users Excel file exported: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting users Excel file: {e}")
            raise
            
    def export_all_bots_stats(self, owner_id: int) -> str:
        """Export all bots statistics for a user"""
        try:
            bots = self.database.get_user_bots_detailed(owner_id)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Botlar Statistikasi"
            
            # Headers
            headers = [
                "Bot nomi", "Username", "Yaratilgan", "Faol", 
                "Konkurslar soni", "Jami ishtirokchilar", "Jami yuborilgan"
            ]
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                
            # Add bot data
            for row, bot in enumerate(bots, 2):
                ws.cell(row=row, column=1, value=bot['name'])
                ws.cell(row=row, column=2, value=f"@{bot['username']}")
                ws.cell(row=row, column=3, value=bot['created_at'])
                ws.cell(row=row, column=4, value="Ha" if bot['active'] else "Yo'q")
                ws.cell(row=row, column=5, value=bot['contests_count'])
                ws.cell(row=row, column=6, value=bot['total_participants'])
                ws.cell(row=row, column=7, value=bot['total_submissions'])
                
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"mening_botlarim_{timestamp}.xlsx"
            
            os.makedirs("exports", exist_ok=True)
            filepath = os.path.join("exports", filename)
            
            wb.save(filepath)
            
            logger.info(f"Bots stats exported: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting bots stats: {e}")
            raise
            
    def export_contest_submissions(self, contest_id: int) -> str:
        """Export all contest submissions with detailed info"""
        try:
            contest = self.database.get_contest_by_id(contest_id)
            submissions = self.database.get_contest_submissions_detailed(contest_id)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Yuborilgan rasmlar"
            
            headers = [
                "№", "Ishtirokchi", "Username", "Yuborilgan vaqti", 
                "File ID", "Caption", "Ovozlar soni"
            ]
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                
            # Add submission data
            for row, submission in enumerate(submissions, 2):
                ws.cell(row=row, column=1, value=row-1)
                ws.cell(row=row, column=2, value=submission['participant_name'])
                ws.cell(row=row, column=3, value=f"@{submission['username']}" if submission['username'] else "N/A")
                ws.cell(row=row, column=4, value=submission['submitted_at'])
                ws.cell(row=row, column=5, value=submission['file_id'])
                ws.cell(row=row, column=6, value=submission['caption'] or "")
                ws.cell(row=row, column=7, value=submission['votes'])
                
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"yuborilgan_rasmlar_{contest['title'].replace(' ', '_')}_{timestamp}.xlsx"
            
            os.makedirs("exports", exist_ok=True)
            filepath = os.path.join("exports", filename)
            
            wb.save(filepath)
            
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting submissions: {e}")
            raise
            
    def export_bot_data(self, bot_id: int) -> str:
        """Export all bot data (participants, contests, submissions, referrals)"""
        try:
            bot_info = self.database.get_bot_by_id(bot_id)
            if not bot_info:
                raise ValueError("Bot not found")
                
            # Get referrals list
            referrals_list = self.database.get_referrals_list(bot_id)
            
            # Create workbook with multiple sheets
            wb = Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)

            # Shared styling definitions
            header_font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            thick_border = Border(
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000')
            )

            data_font = Font(size=14, name="Calibri")
            data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

            light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # Sheet: Bot foydalanuvchilari
            ws_users = wb.create_sheet("Bot foydalanuvchilari")

            if referrals_list:
                # No title header - start directly with data
                user_headers = ["№", "Ism", "Familiya", "Username", "Telegram ID", "Qo'shilgan vaqti"]
                
                # Enhanced styling for headers - larger fonts
                header_font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
                header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Set row height for header - taller
                ws_users.row_dimensions[1].height = 35

                for col, header in enumerate(user_headers, 1):
                    cell = ws_users.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thick_border
                    
                # Enhanced styling for data rows - larger fonts
                data_font = Font(size=14, name="Calibri")
                data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Add referrals data with enhanced formatting and more spacing
                for row, referral in enumerate(referrals_list, 2):
                    # Set row height - taller for more spacing
                    ws_users.row_dimensions[row].height = 28
                    
                    # Choose fill color (alternating rows)
                    fill_color = light_fill if row % 2 == 0 else white_fill
                    
                    # Add data with styling
                    # referral tuple: (first_name, last_name, phone_number, username, telegram_id, created_at)
                    cells = [
                        (row, 1, row - 1),  # №
                        (row, 2, referral[0] or "N/A"),  # first_name
                        (row, 3, referral[1] or "N/A"),  # last_name
                        (row, 4, f"@{referral[3]}" if referral[3] else "N/A"),  # username
                        (row, 5, referral[4]),  # telegram_id
                        (row, 6, referral[5])   # created_at
                    ]
                    
                    for r, c, value in cells:
                        cell = ws_users.cell(row=r, column=c, value=value)
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = thin_border
                        cell.fill = fill_color
                
                # Set optimal column widths for users sheet - wider columns
                users_column_widths = {
                    'A': 10,  # №
                    'B': 25,  # Ism
                    'C': 25,  # Familiya
                    'D': 22,  # Username
                    'E': 18,  # Telegram ID
                    'F': 25   # Qo'shilgan vaqti
                }
                
                for column_letter, width in users_column_widths.items():
                    ws_users.column_dimensions[column_letter].width = width
            else:
                # No users message
                cell = ws_users.cell(row=1, column=1, value="Hozircha bot foydalanuvchilari yo'q")
                cell.font = Font(size=14, name="Calibri")
                ws_users.column_dimensions['A'].width = 50
            
            # Sheet 3: Referral qo'shganlar (Summary by referrer)
            users_with_referrals = self.database.get_users_with_referrals_detailed(bot_id)
            ws3 = wb.create_sheet("Referral qo'shganlar")

            if users_with_referrals:
                summary_headers = [
                    "№", "Ism", "Familiya", "Telefon raqami", "Qo'shganlar soni"
                ]

                ws3.row_dimensions[1].height = 35

                for col, header in enumerate(summary_headers, 1):
                    cell = ws3.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thick_border

                for row, user_data in enumerate(users_with_referrals, 2):
                    ws3.row_dimensions[row].height = 24
                    fill_color = light_fill if row % 2 == 0 else white_fill

                    referral_count = user_data.get('referral_count')
                    if referral_count is None:
                        referred_users = user_data.get('referred_users', [])
                        referral_count = len(referred_users)

                    cells = [
                        (row, 1, row - 1),
                        (row, 2, user_data.get('first_name', 'N/A')),
                        (row, 3, user_data.get('last_name', '')),
                        (row, 4, user_data.get('phone_number', 'Berilmagan')),
                        (row, 5, referral_count)
                    ]

                    for r, c, value in cells:
                        cell = ws3.cell(row=r, column=c, value=value)
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = thin_border
                        cell.fill = fill_color

                summary_column_widths = {
                    'A': 8,
                    'B': 22,
                    'C': 22,
                    'D': 20,
                    'E': 18
                }

                for column_letter, width in summary_column_widths.items():
                    ws3.column_dimensions[column_letter].width = width
            else:
                cell = ws3.cell(row=1, column=1, value="Hozircha referal orqali foydalanuvchi qo'shganlar yo'q")
                cell.font = Font(size=14, name="Calibri")
                ws3.column_dimensions['A'].width = 50
                
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"bot_data_{bot_info['name'].replace(' ', '_')}_{timestamp}.xlsx"
            
            # Ensure exports directory exists
            os.makedirs("exports", exist_ok=True)
            filepath = os.path.join("exports", filename)
            
            # Save file
            wb.save(filepath)
            
            logger.info(f"Bot data exported: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting bot data: {e}")
            raise
            
    def export_referral_statistics(self, bot_id: int) -> str:
        """Export detailed referral statistics with top referrers and all referred users"""
        try:
            bot_info = self.database.get_bot_by_id(bot_id)
            if not bot_info:
                raise ValueError("Bot not found")
                
            # Get top referrers (100 eng ko'p odam qo'shganlar)
            top_referrers = self.database.get_top_referrers(bot_id, limit=100)
            
            # Get all referred users with referrer info
            all_referred_users = self.database.get_all_referred_users_detailed(bot_id)
            
            # Create workbook with multiple sheets
            wb = Workbook()
            
            # Sheet 1: Top 100 Referrers (Eng ko'p odam qo'shganlar)
            ws1 = wb.active
            ws1.title = "TOP 100 Referrerlar"
            
            # Add title
            title_row = self._add_title_to_sheet(ws1, f"🏆 TOP 100 - ENG KO'P ODAM QO'SHGANLAR", 1, 5)
            
            # Headers for top referrers
            headers = [
                "O'rin", "Ism", "Telefon raqami", "Qo'shgan odamlar soni", "Username"
            ]
            
            # Enhanced styling for headers
            header_font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
            header_fill = PatternFill(start_color="D4A017", end_color="D4A017", fill_type="solid")  # Gold color
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Create border style
            thick_border = Border(
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000')
            )
            
            # Set row height for header
            ws1.row_dimensions[title_row].height = 30
            
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=title_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thick_border
            
            # Enhanced styling for data rows
            data_font = Font(size=12, name="Calibri")
            data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Create thin border for data cells
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Special colors for top positions
            gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # 1st place
            silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # 2nd place
            bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")  # 3rd place
            light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            # Add top referrers data
            for row, referrer in enumerate(top_referrers, title_row + 1):
                # Set row height
                ws1.row_dimensions[row].height = 25
                
                # Choose fill color based on position
                position = row - title_row
                if position == 1:
                    fill_color = gold_fill
                    data_font_special = Font(size=12, name="Calibri", bold=True)
                elif position == 2:
                    fill_color = silver_fill
                    data_font_special = Font(size=12, name="Calibri", bold=True)
                elif position == 3:
                    fill_color = bronze_fill
                    data_font_special = Font(size=12, name="Calibri", bold=True)
                else:
                    fill_color = light_fill if position % 2 == 0 else white_fill
                    data_font_special = data_font
                
                # Add medals for top 3
                if position == 1:
                    position_display = "🥇 1"
                elif position == 2:
                    position_display = "🥈 2"
                elif position == 3:
                    position_display = "🥉 3"
                else:
                    position_display = str(position)
                
                # Add data with styling
                cells = [
                    (row, 1, position_display),  # O'rin
                    (row, 2, referrer.get('first_name', 'N/A')),  # Ism
                    (row, 3, referrer.get('phone_number', 'Berilmagan')),  # Telefon raqami
                    (row, 4, referrer.get('referral_count', 0)),  # Qo'shgan odamlar soni
                    (row, 5, f"@{referrer.get('username', 'N/A')}" if referrer.get('username') else "N/A")  # Username
                ]
                
                for r, c, value in cells:
                    cell = ws1.cell(row=r, column=c, value=value)
                    cell.font = data_font_special
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    cell.fill = fill_color
            
            # Set optimal column widths for top referrers
            top_referrer_widths = {
                'A': 12,  # O'rin
                'B': 25,  # Ism
                'C': 20,  # Telefon raqami
                'D': 22,  # Qo'shgan odamlar soni
                'E': 20   # Username
            }
            
            for column_letter, width in top_referrer_widths.items():
                ws1.column_dimensions[column_letter].width = width
            
            # Sheet 2: All Referred Users (Barcha referral orqali kelganlar)
            ws2 = wb.create_sheet("Barcha Referral Kelganlar")
            
            # Add title
            title_row2 = self._add_title_to_sheet(ws2, f"📋 BARCHA REFERRAL ORQALI KELGAN FOYDALANUVCHILAR", 1, 6)
            
            # Headers for all referred users
            referred_headers = [
                "№", "Kelgan odam ismi", "Telefon raqami", "Username", "Qo'shgan odam", "Qo'shilgan vaqti"
            ]
            
            # Enhanced styling for headers
            header_font2 = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
            header_fill2 = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            
            # Set row height for header
            ws2.row_dimensions[title_row2].height = 30
            
            for col, header in enumerate(referred_headers, 1):
                cell = ws2.cell(row=title_row2, column=col, value=header)
                cell.font = header_font2
                cell.fill = header_fill2
                cell.alignment = header_alignment
                cell.border = thick_border
            
            # Add all referred users data
            for row, user in enumerate(all_referred_users, title_row2 + 1):
                # Set row height
                ws2.row_dimensions[row].height = 22
                
                # Alternating row colors
                fill_color = light_fill if (row - title_row2) % 2 == 0 else white_fill
                
                # Add data with styling
                cells = [
                    (row, 1, row - title_row2),  # №
                    (row, 2, user.get('first_name', 'N/A')),  # Kelgan odam ismi
                    (row, 3, user.get('phone_number', 'Berilmagan')),  # Telefon raqami
                    (row, 4, f"@{user.get('username', 'N/A')}" if user.get('username') else "N/A"),  # Username
                    (row, 5, user.get('referrer_name', 'N/A')),  # Qo'shgan odam
                    (row, 6, user.get('joined_at', 'N/A'))  # Qo'shilgan vaqti
                ]
                
                for r, c, value in cells:
                    cell = ws2.cell(row=r, column=c, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    cell.fill = fill_color
            
            # Set optimal column widths for referred users
            referred_widths = {
                'A': 8,   # №
                'B': 25,  # Kelgan odam ismi
                'C': 20,  # Telefon raqami
                'D': 20,  # Username
                'E': 25,  # Qo'shgan odam
                'F': 25   # Qo'shilgan vaqti
            }
            
            for column_letter, width in referred_widths.items():
                ws2.column_dimensions[column_letter].width = width
            
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            bot_name = bot_info.get('name', 'bot').replace(' ', '_')
            filename = f"referral_statistika_{bot_name}_{timestamp}.xlsx"
            
            # Ensure exports directory exists
            os.makedirs("exports", exist_ok=True)
            filepath = os.path.join("exports", filename)
            
            # Save file
            wb.save(filepath)
            
            logger.info(f"Referral statistics exported: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting referral statistics: {e}")
            raise
    
    def export_users_statistics(self) -> str:
        """Export general users statistics across all bots"""
        try:
            # Get all users statistics
            all_users = self.database.get_all_users_for_export()
            bots_stats = self.database.get_all_bots_admin()
            
            # Create workbook
            wb = Workbook()
            
            # Sheet 1: All Users
            ws1 = wb.active
            ws1.title = "Barcha Foydalanuvchilar"
            
            # Add title
            title_row = self._add_title_to_sheet(ws1, f"👥 BARCHA TIZIM FOYDALANUVCHILARI", 1, 6)
            
            # Headers for users
            headers = [
                "№", "Ism", "Familiya", "Username", "Telefon raqami", "Qo'shilgan vaqti"
            ]
            
            # Enhanced styling for headers
            header_font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Create border style
            thick_border = Border(
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000')
            )
            
            # Set row height for header
            ws1.row_dimensions[title_row].height = 30
            
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=title_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thick_border
            
            # Enhanced styling for data rows
            data_font = Font(size=12, name="Calibri")
            data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Create thin border for data cells
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Alternating row colors
            light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            # Add users data
            for row, user in enumerate(all_users, title_row + 1):
                # Set row height
                ws1.row_dimensions[row].height = 22
                
                # Choose fill color (alternating rows)
                fill_color = light_fill if (row - title_row) % 2 == 0 else white_fill
                
                # Add data with styling
                cells = [
                    (row, 1, row - title_row),  # №
                    (row, 2, user.get('first_name', 'N/A')),  # Ism
                    (row, 3, user.get('last_name', 'N/A')),  # Familiya
                    (row, 4, f"@{user.get('username', 'N/A')}" if user.get('username') else "N/A"),  # Username
                    (row, 5, user.get('phone_number', 'Berilmagan')),  # Telefon raqami
                    (row, 6, user.get('created_at', 'N/A'))  # Qo'shilgan vaqti
                ]
                
                for r, c, value in cells:
                    cell = ws1.cell(row=r, column=c, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    cell.fill = fill_color
            
            # Set optimal column widths
            user_widths = {
                'A': 8,   # №
                'B': 25,  # Ism
                'C': 25,  # Familiya
                'D': 20,  # Username
                'E': 20,  # Telefon raqami
                'F': 25   # Qo'shilgan vaqti
            }
            
            for column_letter, width in user_widths.items():
                ws1.column_dimensions[column_letter].width = width
            
            # Sheet 2: Bots Statistics
            ws2 = wb.create_sheet("Botlar Statistikasi")
            
            # Add title
            title_row2 = self._add_title_to_sheet(ws2, f"🤖 BARCHA BOTLAR STATISTIKASI", 1, 7)
            
            # Headers for bots
            bot_headers = [
                "№", "Bot nomi", "Username", "Egasi", "Faol", "Konkurslar", "Yaratilgan"
            ]
            
            # Set row height for header
            ws2.row_dimensions[title_row2].height = 30
            
            for col, header in enumerate(bot_headers, 1):
                cell = ws2.cell(row=title_row2, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thick_border
            
            # Add bots data
            for row, bot in enumerate(bots_stats, title_row2 + 1):
                # Set row height
                ws2.row_dimensions[row].height = 22
                
                # Choose fill color (alternating rows)
                fill_color = light_fill if (row - title_row2) % 2 == 0 else white_fill
                
                # Add data with styling
                cells = [
                    (row, 1, row - title_row2),  # №
                    (row, 2, bot.get('name', 'N/A')),  # Bot nomi
                    (row, 3, f"@{bot.get('username', 'N/A')}" if bot.get('username') else "N/A"),  # Username
                    (row, 4, bot.get('owner_name', 'N/A')),  # Egasi
                    (row, 5, "✅ Faol" if bot.get('active') else "❌ Nofaol"),  # Faol
                    (row, 6, bot.get('contests_count', 0)),  # Konkurslar
                    (row, 7, bot.get('created_at', 'N/A'))  # Yaratilgan
                ]
                
                for r, c, value in cells:
                    cell = ws2.cell(row=r, column=c, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    cell.fill = fill_color
            
            # Set optimal column widths for bots
            bot_widths = {
                'A': 8,   # №
                'B': 25,  # Bot nomi
                'C': 20,  # Username
                'D': 25,  # Egasi
                'E': 15,  # Faol
                'F': 12,  # Konkurslar
                'G': 25   # Yaratilgan
            }
            
            for column_letter, width in bot_widths.items():
                ws2.column_dimensions[column_letter].width = width
            
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"foydalanuvchilar_statistika_{timestamp}.xlsx"
            
            # Ensure exports directory exists
            os.makedirs("exports", exist_ok=True)
            filepath = os.path.join("exports", filename)
            
            # Save file
            wb.save(filepath)
            
            logger.info(f"Users statistics exported: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting users statistics: {e}")
            raise
    
    def export_users_only(self, bot_id: int) -> str:
        """Export only bot users (foydalanuvchilar) with referral counts to Excel"""
        try:
            bot_info = self.database.get_bot_by_id(bot_id)
            if not bot_info:
                raise ValueError("Bot not found")
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Bot foydalanuvchilari"
            
            # Get all users with referral counts
            users_list = self.database.get_all_users_with_referrals_for_export(bot_id)
            
            # Styling
            header_font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            thick_border = Border(
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000')
            )
            
            data_font = Font(size=14, name="Calibri")
            data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            if users_list:
                # Headers
                headers = ["№", "Ism", "Familiya", "Telefon raqami", "Referral qo'shganlar soni"]
                
                ws.row_dimensions[1].height = 35
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thick_border
                
                # Add data
                for row, user in enumerate(users_list, 2):
                    ws.row_dimensions[row].height = 28
                    fill_color = light_fill if row % 2 == 0 else white_fill
                    
                    cells = [
                        (row, 1, row - 1),
                        (row, 2, user.get('first_name', 'N/A')),
                        (row, 3, user.get('last_name', 'N/A')),
                        (row, 4, user.get('phone_number', 'Berilmagan')),
                        (row, 5, user.get('referral_count', 0))
                    ]
                    
                    for r, c, value in cells:
                        cell = ws.cell(row=r, column=c, value=value)
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = thin_border
                        cell.fill = fill_color
                
                # Set column widths
                column_widths = {
                    'A': 10,  # №
                    'B': 25,  # Ism
                    'C': 25,  # Familiya
                    'D': 22,  # Telefon raqami
                    'E': 28   # Referral qo'shganlar soni
                }
                
                for column_letter, width in column_widths.items():
                    ws.column_dimensions[column_letter].width = width
            else:
                cell = ws.cell(row=1, column=1, value="Hozircha bot foydalanuvchilari yo'q")
                cell.font = Font(size=14, name="Calibri")
                ws.column_dimensions['A'].width = 50
            
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{bot_info['name']}_foydalanuvchilar_{timestamp}.xlsx"
            
            # Ensure exports directory exists
            os.makedirs("exports", exist_ok=True)
            filepath = os.path.join("exports", filename)
            
            # Save file
            wb.save(filepath)
            
            logger.info(f"Bot users exported: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting bot users: {e}")
            raise
    
    def export_referrals_only(self, bot_id: int) -> str:
        """Export only referral data to Excel"""
        try:
            bot_info = self.database.get_bot_by_id(bot_id)
            if not bot_info:
                raise ValueError("Bot not found")
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Referral qo'shganlar"
            
            # Get users with referrals
            users_with_referrals = self.database.get_users_with_referrals_detailed(bot_id)
            
            # Styling
            header_font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            thick_border = Border(
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000')
            )
            
            data_font = Font(size=14, name="Calibri")
            data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            if users_with_referrals:
                # Headers
                headers = ["№", "Ism", "Familiya", "Telefon", "Username", "Qo'shganlar soni"]
                
                ws.row_dimensions[1].height = 35
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thick_border
                
                # Add data
                for row, user_data in enumerate(users_with_referrals, 2):
                    ws.row_dimensions[row].height = 28
                    fill_color = light_fill if row % 2 == 0 else white_fill
                    
                    referral_count = user_data.get('referral_count')
                    if referral_count is None:
                        referred_users = user_data.get('referred_users', [])
                        referral_count = len(referred_users)
                    
                    cells = [
                        (row, 1, row - 1),
                        (row, 2, user_data.get('first_name', 'N/A')),
                        (row, 3, user_data.get('last_name', '')),
                        (row, 4, user_data.get('phone_number', 'Berilmagan')),
                        (row, 5, f"@{user_data.get('username', '')}" if user_data.get('username') else "N/A"),
                        (row, 6, referral_count)
                    ]
                    
                    for r, c, value in cells:
                        cell = ws.cell(row=r, column=c, value=value)
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = thin_border
                        cell.fill = fill_color
                
                # Set column widths
                column_widths = {
                    'A': 10,  # №
                    'B': 25,  # Ism
                    'C': 25,  # Familiya
                    'D': 20,  # Telefon
                    'E': 22,  # Username
                    'F': 20   # Qo'shganlar soni
                }
                
                for column_letter, width in column_widths.items():
                    ws.column_dimensions[column_letter].width = width
            else:
                cell = ws.cell(row=1, column=1, value="Hozircha referral qo'shganlar yo'q")
                cell.font = Font(size=14, name="Calibri")
                ws.column_dimensions['A'].width = 50
            
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{bot_info['name']}_referrallar_{timestamp}.xlsx"
            
            # Ensure exports directory exists
            os.makedirs("exports", exist_ok=True)
            filepath = os.path.join("exports", filename)
            
            # Save file
            wb.save(filepath)
            
            logger.info(f"Referrals exported: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting referrals: {e}")
            raise